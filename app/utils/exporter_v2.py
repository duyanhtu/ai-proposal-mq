import json
import os
import shutil
import tempfile
from datetime import datetime
from typing import Any, Dict, List

import openpyxl
from fastapi import HTTPException
from fastapi.responses import FileResponse
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, Side

from app.storage import pgdb

BASE_DIR = os.environ.get('APP_BASE_DIR', os.path.dirname(
    os.path.dirname(os.path.dirname(__file__))))
# Hàm tìm vị trí header và dữ liệu tương ứng


def find_headers(sheet, targets: List[str]) -> List[Dict[str, Any]]:
    """
    Tìm vị trí của các header trong sheet
    """
    arr_targets = [t.strip().lower() for t in targets]
    header_positions = []
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell) and cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value in arr_targets:
                    data_key = {
                        "nhân sự": "personal_requirement",
                        "năng lực tài chính": "finance_requirement",
                        "năng lực kinh nghiệm": "experience_requirement",
                    }.get(cell_value)
                    header_positions.append(
                        {
                            "value": cell_value,
                            "stt": sheet.cell(row=cell.row, column=1).value,
                            "cell": cell,
                            "data_key": data_key,
                        }
                    )
    return sorted(header_positions, key=lambda x: x["cell"].row, reverse=True)


# Hàm gộp description theo position
def merge_descriptions_by_position(data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    merge_descriptions_by_position: Gộp các mô tả theo position
    """
    if not data:
        return []

    # Tạo dictionary để nhóm theo position
    grouped_data = {}
    for item in data:
        position = item.get("position", "")
        if position not in grouped_data:
            grouped_data[position] = {
                "position": position,
                "proposal_id": item.get("proposal_id", ""),
                "quantity": item.get("quantity", ""),
                "document_name": item.get("document_name", ""),
                "descriptions": [],
            }
        grouped_data[position]["descriptions"].append(
            {"name": item.get("name", ""),
             "description": item.get("description", "")}
        )
    # print("grouped_data: ", grouped_data)
    # Chuyển đổi thành list và gộp description
    merged_data = []
    for position, details in grouped_data.items():
        combined_description = "\n".join(
            f"- {desc['name']}: {desc['description']}."
            for desc in details["descriptions"]
        )
        merged_item = {
            "position": position,
            "proposal_id": details["proposal_id"],
            "quantity": details["quantity"],
            "document_name": details["document_name"],
            "description": combined_description,
        }
        merged_data.append(merged_item)
    # print("Merged_data: ", merged_data)
    return merged_data


# Hàm điền dữ liệu vào sheet và format lại
def fill_data(sheet, header_position, data: List[Dict[str, Any]]):
    """
    Điền dữ liệu vào sheet Excel
    """
    start_row = header_position["cell"].row + 1
    length_data = len(data)
    sheet.insert_rows(start_row, amount=length_data)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    for idx, item in enumerate(data):
        row_num = start_row + idx
        is_personal = header_position["value"] == "nhân sự"

        # Lấy giá trị và thay None bằng chuỗi rỗng
        stt_value = f"{header_position['stt']}.{idx + 1}"
        position_or_requirements = (
            f"{str(item.get("position" if is_personal else "requirements", "") or "")}"
        )
        description = f"{str(item.get("description", "") or "")}"
        quantity = f"- Số lượng: {str(item.get("quantity", "") or "")}"
        document_name = f"{str(item.get("document_name", "") or "")}"

        cell_a = sheet.cell(row=row_num, column=1)
        cell_a.value = f"{stt_value}"
        cell_a.alignment = center_alignment
        cell_a.font = Font(bold=True)

        cell_b = sheet.cell(row=row_num, column=2)
        cell_b.value = position_or_requirements
        cell_b.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        cell_b.font = Font(bold=True)

        cell_c = sheet.cell(row=row_num, column=3)
        if is_personal:
            cell_c.value = (
                f"""'{quantity}\n{description}."""
                # f"{item.get('name', '')}: {item.get('description', '')}.\n"
                # f"{description}.\n"
            )
        else:
            cell_c.value = f"{description}."
        cell_c.font = Font(bold=False)
        cell_c.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

        cell_d = sheet.cell(row=row_num, column=4)
        cell_d.value = "x"
        cell_d.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell_d.font = Font(bold=False)

        cell_e = sheet.cell(row=row_num, column=5)
        cell_e.value = document_name
        cell_e.font = Font(bold=True)
        cell_e.alignment = center_alignment

        sheet.row_dimensions[row_num].height = max(
            100, 15 * (cell_c.value or "").count("\n") + 35
        )

        # Áp dụng viền cho toàn bộ dòng
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row_num, column=col).border = thin_border


def process_excel_file_no_upload(id: int):
    """
    Xử lý file Excel và trả về đường dẫn file đã xử lý
    """
    try:
        # Check for valid ID
        if not isinstance(id, int) or id <= 0:
            raise ValueError("ID phải là một số nguyên dương")
            
        template_file_path = os.path.join(
            BASE_DIR,
            "temp",
            "template_checklist_v1.xlsx",

        )
        
        # Verify template file exists
        if not os.path.exists(template_file_path):
            raise FileNotFoundError(f"File template không tồn tại tại đường dẫn: {template_file_path}")
            
        try:
            # Database operations
            proposal_sql = f"""select * from proposal where id = {id}"""
            proposal = pgdb.select(proposal_sql)
            if not proposal:
                raise HTTPException(
                    status_code=404, detail="Không tìm thấy hồ sơ thầu với ID này"
                )

            personal_requirement_sql = f"""
                select hr.position, hr.proposal_id , hr.quantity, hdr.name, hdr.description, hdr.document_name
                from hr_requirement hr, hr_detail_requirement hdr
                where hr.proposal_id = {id} and hr.id = hdr.hr_id
            """
            experience_requirement_sql = (
                f"""select * from experience_requirement where proposal_id = {id}"""
            )
            finance_requirement_sql = (
                f"""select * from finance_requirement where proposal_id = {id}"""
            )

            data_map = {
                "personal_requirement": pgdb.select(personal_requirement_sql),
                "experience_requirement": pgdb.select(experience_requirement_sql),
                "finance_requirement": pgdb.select(finance_requirement_sql),
            }
        except Exception as db_error:
            raise HTTPException(
                status_code=500, detail=f"Lỗi truy vấn cơ sở dữ liệu: {str(db_error)}"
            ) from db_error

        # Gộp description theo position cho personal_requirement
        if "personal_requirement" in data_map and data_map["personal_requirement"]:
            try:
                data_map["personal_requirement"] = merge_descriptions_by_position(
                    data_map["personal_requirement"]
                )
            except Exception as merge_error:
                raise HTTPException(
                    status_code=500, detail=f"Lỗi khi gộp dữ liệu nhân sự: {str(merge_error)}"
                ) from merge_error

        # Tạo thư mục tạm để lưu file nếu chưa tồn tại
        temp_file_path = None
        try:
            timestamp = datetime.now().strftime("%Y_%m_%d_%S_%M_%H")
            with tempfile.NamedTemporaryFile(suffix=".xlsx", prefix=f"Checklist_HSMT_{timestamp}", delete=False) as temp_file:
                temp_file_path = temp_file.name
                # Sao chép file template vào file tạm
                shutil.copyfile(template_file_path, temp_file_path)
        except Exception as temp_file_error:
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass
            raise HTTPException(
                status_code=500, detail=f"Lỗi khi tạo file tạm: {str(temp_file_error)}"
            ) from temp_file_error

        try:
            # Đọc file Excel từ UploadFile
            workbook = openpyxl.load_workbook(temp_file_path)

            # Xử lý file Excel
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                if (
                    worksheet["A1"].value
                    and "CHECKLIST HỒ SƠ MỜI THẦU" in str(worksheet["A1"].value).upper()
                ):
                    worksheet["A1"].value = f"{worksheet['A1'].value.strip()} ({id})"

            # Các trường cần điền giá trị
            fields_to_find = {
                "selection_method": str(proposal[0]["selection_method"] or ""),
                "proposal_name": str(proposal[0]["proposal_name"] or ""),
                "investor_name": str(proposal[0]["investor_name"] or ""),
                "release_date": str(proposal[0]["release_date"] or ""),
                "project": str(proposal[0]["project"] or ""),
                "closing_time": str(proposal[0]["closing_time"] or ""),
            }

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for field_name, field_value in fields_to_find.items():
                    target_row = None
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and field_name.lower() in str(cell.value).lower():
                                target_row = cell.row
                                cell.value = ""
                                break
                        if target_row:
                            break

                    if target_row:
                        col_a_cell = sheet.cell(row=target_row, column=1)
                        if col_a_cell.value:
                            new_value = f"{col_a_cell.value.strip()} : {field_value}"
                            col_a_cell.value = new_value

            # Xử lý dữ liệu cần điền vào bảng
            targets = ["Năng lực tài chính", "Năng lực kinh nghiệm", "Nhân sự"]
            header_positions = find_headers(sheet, targets)
            for header in header_positions:
                data = data_map[header["data_key"]]
                if data:
                    fill_data(sheet, header, data)

            # Lưu lại file Excel đã xử lý
            workbook.save(temp_file_path)
            # Tạo metadata
            file_name_export = os.path.basename(temp_file_path)
            metadata = {"status": "Success!", "file_name": file_name_export}
            metadata_json = json.dumps(metadata)

            # Trả về file Excel đã chỉnh sửa
            return FileResponse(
                path=temp_file_path,
                filename=file_name_export,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"X-Metadata": metadata_json},
            )

        except Exception as excel_error:
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass
            raise HTTPException(
                status_code=500, detail=f"Xử lý file Excel thất bại: {str(excel_error)}"
            ) from excel_error
            
    except HTTPException:
        # Re-raise HTTP exceptions without modification
        raise
    except Exception as e:
        # Catch any other unexpected errors
        raise HTTPException(
            status_code=500, detail=f"Lỗi không xác định: {str(e)}"
        ) from e